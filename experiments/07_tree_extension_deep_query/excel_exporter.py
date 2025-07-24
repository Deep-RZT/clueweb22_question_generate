#!/usr/bin/env python3
"""
优化版Excel导出器 - 专门处理dict格式的JSON数据
支持嵌套累积型和LLM整合型综合问题及答案的完整导出
"""

import json
import pandas as pd
from pathlib import Path
from typing import Dict, List, Any, Optional
import logging

logger = logging.getLogger(__name__)

class FixedCleanExcelExporter:
    """优化版Excel导出器 - 完全dict格式支持"""
    
    def __init__(self, output_dir: str = "results"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)

    def export_clean_excel(self, json_file_path: Path) -> str:
        """导出简洁版Excel文件"""
        print(f"🔄 读取JSON文件: {json_file_path.name}")
        
        # 读取JSON数据
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        print("📊 解析数据...")
        parsed_data = self._parse_dict_data(data)
        
        # 生成Excel文件名
        timestamp = json_file_path.stem.split('_')[-1]
        excel_filename = f"agent_reasoning_analysis_{timestamp}.xlsx"
        excel_path = self.output_dir / excel_filename
        
        print("📝 生成Excel工作表...")
        self._write_excel_sheets(parsed_data, excel_path)
        
        return str(excel_path)

    def _parse_dict_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """解析dict格式的JSON数据"""
        parsed = {
            'efficiency_data': [],   # 文档处理效率统计
            'composite_qa': [],      # 糅合后的问答对
            'all_process_qa': [],    # 所有过程中的问答对
            'trajectories': [],      # 轨迹数据
        }
        
        processing_results = data.get('processing_results', {})
        processed_docs = processing_results.get('processed_documents', [])
        
        print(f"📋 解析 {len(processed_docs)} 个处理文档...")
        
        for doc_idx, doc in enumerate(processed_docs):
            doc_id = doc.get('doc_id', f'Unknown_{doc_idx}')
            reasoning_trees = doc.get('reasoning_trees', [])
            trajectory_records = doc.get('trajectory_records', [])
            
            print(f"  📄 处理文档: {doc_id} ({len(reasoning_trees)} 推理树)")
            
            # 解析推理树（现在应该是dict格式）
            for tree_idx, tree_data in enumerate(reasoning_trees):
                if isinstance(tree_data, dict):  # dict格式数据
                    tree_id = tree_data.get('tree_id', f'{doc_id}_tree_{tree_idx}')
                    final_composite = tree_data.get('final_composite_query', {})
                    
                    # 获取根节点答案
                    root_node = tree_data.get('root_node', {})
                    root_query = root_node.get('query', {})
                    root_answer = root_query.get('answer', 'N/A')
                    
                    # 处理三格式的综合问题和答案
                    if isinstance(final_composite, dict):
                        # 新的三格式（支持问题、答案和兜底标记）
                        nested_question = final_composite.get('nested_cumulative', '')
                        nested_answer = final_composite.get('nested_cumulative_answer', root_answer)
                        nested_fallback = final_composite.get('nested_cumulative_fallback', False)
                        
                        llm_question = final_composite.get('llm_integrated', '')
                        llm_answer = final_composite.get('llm_integrated_answer', root_answer)
                        llm_fallback = final_composite.get('llm_integrated_fallback', False)
                        
                        ambiguous_question = final_composite.get('ambiguous_integrated', '')
                        ambiguous_answer = final_composite.get('ambiguous_integrated_answer', root_answer)
                        ambiguous_fallback = final_composite.get('ambiguous_integrated_fallback', False)
                        
                        # 检查三种格式是否有效
                        nested_valid = nested_question and len(nested_question.strip()) > 30
                        llm_valid = llm_question and len(llm_question.strip()) > 30
                        ambiguous_valid = ambiguous_question and len(ambiguous_question.strip()) > 30
                        
                        # 添加嵌套累积型
                        parsed['composite_qa'].append({
                            'doc_id': doc_id,
                            'tree_id': tree_id,
                            'composite_question': nested_question if nested_valid else '❌ 嵌套累积型问题生成失败',
                            'composite_answer': nested_answer if nested_valid else '❌ 嵌套累积型答案生成失败',
                            'target_answer': root_answer,
                            'question_length': len(nested_question) if nested_valid else 0,
                            'tree_index': tree_idx,
                            'question_type': '嵌套累积型',
                            'is_valid': nested_valid,
                            'is_fallback': nested_fallback
                        })
                        
                        # 添加LLM整合型
                        parsed['composite_qa'].append({
                            'doc_id': doc_id,
                            'tree_id': tree_id,
                            'composite_question': llm_question if llm_valid else '❌ LLM整合型问题生成失败',
                            'composite_answer': llm_answer if llm_valid else '❌ LLM整合型答案生成失败',
                            'target_answer': root_answer,
                            'question_length': len(llm_question) if llm_valid else 0,
                            'tree_index': tree_idx,
                            'question_type': 'LLM整合型',
                            'is_valid': llm_valid,
                            'is_fallback': llm_fallback
                        })
                        
                        # 添加模糊化整合型
                        parsed['composite_qa'].append({
                            'doc_id': doc_id,
                            'tree_id': tree_id,
                            'composite_question': ambiguous_question if ambiguous_valid else '❌ 模糊化整合型问题生成失败',
                            'composite_answer': ambiguous_answer if ambiguous_valid else '❌ 模糊化整合型答案生成失败',
                            'target_answer': root_answer,
                            'question_length': len(ambiguous_question) if ambiguous_valid else 0,
                            'tree_index': tree_idx,
                            'question_type': '模糊化整合型',
                            'is_valid': ambiguous_valid,
                            'is_fallback': ambiguous_fallback
                        })
                    else:
                        # 兼容旧格式
                        composite_question = str(final_composite) if final_composite else ''
                        is_valid = len(composite_question.strip()) > 30
                        
                        parsed['composite_qa'].append({
                            'doc_id': doc_id,
                            'tree_id': tree_id,
                            'composite_question': composite_question if is_valid else '❌ 未生成有效综合问题',
                            'composite_answer': root_answer,
                            'target_answer': root_answer,
                            'question_length': len(composite_question) if is_valid else 0,
                            'tree_index': tree_idx,
                            'question_type': '旧格式（单一类型）',
                            'is_valid': is_valid,
                            'is_fallback': True  # 旧格式都标记为兜底
                        })
                    
                    # 提取所有层级的问答对 - dict格式
                    process_qa = self._extract_qa_from_dict_tree(tree_data, doc_id, tree_id, tree_idx)
                    parsed['all_process_qa'].extend(process_qa)
                    
                else:
                    # 字符串格式数据 - 不应该出现在新系统中
                    print(f"    ⚠️ 发现字符串格式推理树，跳过处理")
                    continue
            
            # 解析轨迹记录
            for traj in trajectory_records:
                if isinstance(traj, dict):
                    traj_info = self._parse_trajectory(traj, doc_id)
                    if traj_info:
                        parsed['trajectories'].append(traj_info)
            
            # 生成文档级效率数据
            valid_composites = sum(1 for comp in parsed['composite_qa'] if comp['doc_id'] == doc_id and comp['is_valid'])
            doc_efficiency = {
                'doc_id': doc_id,
                'processing_time': doc.get('processing_time', 0),
                'trees_generated': len(reasoning_trees),
                'valid_composite_questions': valid_composites,
                'invalid_composite_questions': len(reasoning_trees) * 2 - valid_composites,  # 乘以2因为每个树生成2种类型
                'success': len(reasoning_trees) > 0
            }
            parsed['efficiency_data'].append(doc_efficiency)
        
        return parsed

    def _extract_qa_from_dict_tree(self, tree_data: Dict[str, Any], doc_id: str, tree_id: str, tree_idx: int) -> List[Dict[str, Any]]:
        """从dict格式的推理树中提取所有问答对"""
        qa_pairs = []
        
        try:
            all_nodes = tree_data.get('all_nodes', {})
            
            for node_id, node_data in all_nodes.items():
                query_data = node_data.get('query', {})
                
                qa_pairs.append({
                    'doc_id': doc_id,
                    'tree_id': tree_id,
                    'tree_index': tree_idx,
                    'node_id': node_id,
                    'question': query_data.get('query_text', 'N/A'),
                    'answer': query_data.get('answer', 'N/A'),
                    'branch_type': node_data.get('branch_type', 'unknown'),
                    'layer': node_data.get('layer', 0),
                    'generation_method': query_data.get('generation_method', 'unknown'),
                    'validation_passed': query_data.get('validation_passed', False),
                    'minimal_keywords': len(query_data.get('minimal_keywords', [])),
                    'parent_query_id': query_data.get('parent_query_id'),
                    'extension_type': query_data.get('extension_type', 'unknown')
                })
            
            print(f"    ✅ 提取了 {len(qa_pairs)} 个问答对")
            
        except Exception as e:
            print(f"    ❌ 提取问答对失败: {e}")
            logger.error(f"从树 {tree_id} 提取问答对失败: {e}")
        
        return qa_pairs

    def _parse_trajectory(self, traj: Dict, doc_id: str) -> Optional[Dict]:
        """解析轨迹记录"""
        try:
            # 从实际的轨迹记录字段中提取数据
            keywords_list = traj.get('current_keywords', [])
            keywords_str = ', '.join(keywords_list) if isinstance(keywords_list, list) else str(keywords_list)
            
            # 获取验证结果
            validation_results = traj.get('validation_results', {})
            validation_passed = validation_results.get('validation_passed', False) if isinstance(validation_results, dict) else False
            
            return {
                'doc_id': doc_id,
                'step': traj.get('step', 'unknown'),
                'layer_level': traj.get('layer_level', 0),
                'query_id': traj.get('query_id', 'N/A'),
                'query_text': traj.get('current_question', 'N/A')[:200],
                'answer': traj.get('current_answer', 'N/A')[:200],
                'minimal_keywords': keywords_str[:100],  # 限制长度
                'generation_method': traj.get('generation_method', 'unknown'),
                'validation_passed': validation_passed,
                'processing_time_ms': traj.get('processing_time_ms', 0),
                'extension_type': traj.get('extension_type', 'N/A'),
                'tree_id': traj.get('tree_id', 'N/A'),
                'parent_question': traj.get('parent_question', 'N/A')[:100] if traj.get('parent_question') else 'N/A',
                'parent_answer': traj.get('parent_answer', 'N/A')[:100] if traj.get('parent_answer') else 'N/A',
                'circular_check': traj.get('circular_check_result', 'N/A'),
                'api_calls': traj.get('api_calls_count', 0)
            }
        except Exception as e:
            print(f"解析轨迹记录失败: {e}")
            return None

    def _write_excel_sheets(self, parsed_data: Dict[str, Any], excel_path: Path):
        """写入Excel工作表"""
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            
            # Sheet1: 文档处理效率统计
            self._write_efficiency_data(parsed_data['efficiency_data'], writer)
            
            # Sheet2: 所有过程中的问答对
            self._write_all_process_qa(parsed_data['all_process_qa'], writer)
            
            # Sheet3: 推理轨迹记录
            self._write_trajectories(parsed_data['trajectories'], writer)
            
            # Sheet4: 糅合后的综合问答（三格式+兜底标记）
            self._write_final_composite_qa(parsed_data['composite_qa'], writer)
        
        print(f"✅ Excel文件已生成: {excel_path}")

    def _write_efficiency_data(self, efficiency_data: List[Dict], writer):
        """写入文档处理效率统计"""
        if not efficiency_data:
            return
        
        df = pd.DataFrame(efficiency_data)
        df = df.reindex(columns=[
            'doc_id', 'processing_time', 'trees_generated', 
            'valid_composite_questions', 'invalid_composite_questions', 'success'
        ])
        
        # 重命名列
        df.columns = ['文档ID', '处理时间(秒)', '生成推理树数', '有效综合问题数', '无效问题数', '处理成功']
        
        df.to_excel(writer, sheet_name='Sheet1-文档处理效率统计', index=False)

    def _write_all_process_qa(self, all_process_qa: List[Dict], writer):
        """写入所有过程中的问答对"""
        if not all_process_qa:
            return
        
        df = pd.DataFrame(all_process_qa)
        df = df.reindex(columns=[
            'doc_id', 'tree_id', 'node_id', 'layer', 'branch_type', 
            'question', 'answer', 'generation_method', 'validation_passed'
        ])
        
        # 重命名列
        df.columns = [
            '文档ID', '推理树ID', '节点ID', '层级', '分支类型', 
            '问题', '答案', '生成方法', '验证通过'
        ]
        
        df.to_excel(writer, sheet_name='Sheet2-所有过程中的问答对', index=False)

    def _write_trajectories(self, trajectories: List[Dict], writer):
        """写入推理轨迹记录"""
        if not trajectories:
            return
        
        df = pd.DataFrame(trajectories)
        # 更新列定义以匹配新的轨迹记录字段
        available_columns = [col for col in [
            'doc_id', 'step', 'layer_level', 'query_id', 'query_text', 'answer', 
            'minimal_keywords', 'generation_method', 'validation_passed', 
            'processing_time_ms', 'extension_type', 'tree_id', 'parent_question', 
            'parent_answer', 'circular_check', 'api_calls'
        ] if col in df.columns]
        
        df = df.reindex(columns=available_columns)
        
        # 动态重命名列，以匹配实际可用的列
        column_mapping = {
            'doc_id': '文档ID',
            'step': '步骤',
            'layer_level': '层级',
            'query_id': '查询ID',
            'query_text': '查询文本',
            'answer': '答案',
            'minimal_keywords': '最小关键词',
            'generation_method': '生成方法',
            'validation_passed': '验证通过',
            'processing_time_ms': '处理时间(ms)',
            'extension_type': '扩展类型',
            'tree_id': '推理树ID',
            'parent_question': '父问题',
            'parent_answer': '父答案',
            'circular_check': '循环检查',
            'api_calls': 'API调用次数'
        }
        
        # 只重命名存在的列
        new_column_names = [column_mapping.get(col, col) for col in df.columns]
        df.columns = new_column_names
        
        df.to_excel(writer, sheet_name='Sheet3-推理轨迹记录', index=False)

    def _write_final_composite_qa(self, composite_qa: List[Dict], writer):
        """写入糅合后的综合问答（三格式）"""
        if not composite_qa:
            return
        
        # 准备数据
        composite_data = []
        
        for idx, comp in enumerate(composite_qa):
            question_type = comp['question_type']
            status = "✅ 有效" if comp['is_valid'] else "❌ 无效"
            fallback_status = "🔄 兜底" if comp.get('is_fallback', False) else "✅ 生产"
            
            composite_data.append({
                '序号': idx + 1,
                '文档ID': comp['doc_id'],
                '推理树ID': comp['tree_id'],
                '问题类型': question_type,
                '糅合问题': comp['composite_question'],
                '糅合答案': comp['composite_answer'],
                '最终答案': comp['target_answer'],
                '问题状态': status,
                '生成方式': fallback_status,
                '问题长度': comp['question_length'],
                '树索引': comp['tree_index']
            })
        
        df = pd.DataFrame(composite_data)
        
        # 应用Excel格式化，兜底结果用不同颜色标记
        df.to_excel(writer, sheet_name='Sheet4-糅合后的综合问答', index=False)
        
        # 获取工作表以应用格式化
        workbook = writer.book
        worksheet = writer.sheets['Sheet4-糅合后的综合问答']
        
        # 导入openpyxl样式
        from openpyxl.styles import PatternFill, Font
        
        # 定义样式
        fallback_fill = PatternFill(start_color='FFE6CC', end_color='FFE6CC', fill_type='solid')  # 橙色背景
        production_fill = PatternFill(start_color='E6F3E6', end_color='E6F3E6', fill_type='solid')  # 绿色背景
        
        # 应用条件格式
        for row_idx, row in enumerate(df.itertuples(), start=2):  # 从第2行开始（跳过表头）
            generation_method = row[9]  # '生成方式'列
            if '兜底' in generation_method:
                # 对整行应用兜底样式
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = fallback_fill
            elif '生产' in generation_method:
                # 对整行应用生产样式
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = production_fill 