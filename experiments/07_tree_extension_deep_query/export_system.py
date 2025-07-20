"""
Export System for Tree Extension Deep Query Results
Generates JSON and Excel formatted results for comprehensive analysis
"""

import logging
import json
import pandas as pd
from typing import List, Dict, Any, Optional
from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import re # Added for _extract_field

# Imports for type hints - loaded dynamically when needed

# 配置日志
logger = logging.getLogger(__name__)

class ExportSystem:
    """Export system for comprehensive results output"""
    
    def __init__(self):
        # Get configuration
        from config import get_config
        self.config = get_config()
        
        self.results_dir = Path(self.config.output_dir)
        self.export_formats = self.config.export_formats
        self.include_trajectories = self.config.record_full_trajectory
        
        # Ensure results directory exists
        self.results_dir.mkdir(exist_ok=True)
        
        # Statistics
        self.stats = {
            'total_exports': 0,
            'json_exports': 0,
            'excel_exports': 0,
            'failed_exports': 0
        }
    
    def export_experiment_results(self, question_trees: List[Any], 
                                trajectories: List[Any],
                                experiment_metadata: Dict[str, Any]) -> Dict[str, str]:
        """导出完整的实验结果"""
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        export_results = {}
        
        logger.info(f"开始导出实验结果 - 时间戳: {timestamp}")
        
        try:
            # 准备导出数据
            export_data = self._prepare_export_data(question_trees, trajectories, experiment_metadata)
            
            # 导出JSON格式
            if 'json' in self.export_formats:
                json_file = self._export_json(export_data, timestamp)
                export_results['json'] = str(json_file)
                self.stats['json_exports'] += 1
            
            # 导出Excel格式
            if 'excel' in self.export_formats:
                excel_file = self._export_excel(export_data, timestamp)
                if excel_file: # Only add to results if export was successful
                    export_results['excel'] = str(excel_file)
                    self.stats['excel_exports'] += 1
            
            # 生成分析报告
            if self.config.include_analysis_report:
                analysis_file = self._generate_analysis_report(export_data, timestamp)
                export_results['analysis'] = str(analysis_file)
            
            self.stats['total_exports'] += 1
            logger.info(f"实验结果导出完成: {len(export_results)} 个文件")
            
            return export_results
            
        except Exception as e:
            logger.error(f"导出实验结果失败: {e}")
            self.stats['failed_exports'] += 1
            return {}
    
    def _prepare_export_data(self, question_trees: List[Any],
                           trajectories: List[Any],
                           experiment_metadata: Dict[str, Any]) -> Dict[str, Any]:
        """准备导出数据"""
        
        # 准备问题树数据
        trees_data = []
        for tree in question_trees:
            tree_data = {
                'tree_id': f"tree_{len(trees_data) + 1}",
                'root_question': getattr(tree, 'root_question', {}),
                'nodes': getattr(tree, 'nodes', {}),
                'total_nodes': getattr(tree, 'total_nodes', 0),
                'max_depth': getattr(tree, 'current_depth', 0),
                'extensions_count': len(getattr(tree, 'nodes', {}))
            }
            trees_data.append(tree_data)
        
        # 准备轨迹数据
        trajectories_data = []
        for trajectory in trajectories:
            traj_data = {
                'trajectory_id': getattr(trajectory, 'trajectory_id', ''),
                'document_id': getattr(trajectory, 'document_id', ''),
                'total_hops': getattr(trajectory, 'total_hops', 0),
                'success_rate': getattr(trajectory, 'success_rate', 0.0),
                'steps': getattr(trajectory, 'extension_steps', [])
            }
            trajectories_data.append(traj_data)
        
        return {
            'experiment_metadata': experiment_metadata,
            'question_trees': trees_data,
            'trajectories': trajectories_data,
            'export_timestamp': datetime.now().isoformat(),
            'statistics': {
                'total_trees': len(question_trees),
                'total_trajectories': len(trajectories),
                'total_extensions': sum(tree_data['extensions_count'] for tree_data in trees_data)
            }
        }
    
    def _extract_tree_structure(self, tree: Any) -> Dict[str, Any]:
        """提取树结构"""
        structure = {
            'root': {
                'question': tree.root_question.text,
                'answer': tree.root_question.short_answer,
                'type': 'root'
            },
            'nodes': {}
        }
        
        for node_id, node in tree.nodes.items():
            structure['nodes'][node_id] = {
                'question': node.question,
                'answer': node.answer,
                'short_answer': getattr(node, 'short_answer', None),
                'type': node.extension_type,
                'depth': node.depth_level,
                'parent': node.parent_node_id,
                'confidence': node.confidence,
                'verified': node.search_verified
            }
        
        return structure
    
    def _export_json(self, export_data: Dict[str, Any], timestamp: str) -> Path:
        """导出JSON格式"""
        json_file = self.results_dir / f"tree_extension_experiment_{timestamp}.json"
        
        try:
            with open(json_file, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, indent=2, ensure_ascii=False, default=str)
            
            logger.info(f"JSON结果已导出: {json_file}")
            return json_file
            
        except Exception as e:
            logger.error(f"JSON导出失败: {e}")
            raise
    
    def _export_excel(self, export_data: Dict[str, Any], timestamp: str) -> Path:
        """导出Excel格式 - 重写版本"""
        excel_file = self.results_dir / f"tree_extension_experiment_{timestamp}.xlsx"
        
        try:
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                # 1. 主要问答表
                qa_pairs = []
                for tree_data in export_data.get('question_trees', []):
                    tree_id = tree_data.get('tree_id', 'unknown')
                    
                    # 直接从对象结构提取根问题信息
                    root_question = tree_data.get('root_question', {})
                    
                    # 处理根问题 - 可能是对象或字符串
                    if hasattr(root_question, 'question_text'):
                        # BaseQuestion对象结构
                        question_text = root_question.question_text
                        document_id = root_question.document_id
                        if hasattr(root_question, 'short_answer') and hasattr(root_question.short_answer, 'answer_text'):
                            short_answer_text = root_question.short_answer.answer_text
                        else:
                            short_answer_text = ''
                    elif isinstance(root_question, dict):
                        # 字典结构
                        question_text = root_question.get('question_text', '')
                        document_id = root_question.get('document_id', '')
                        short_answer = root_question.get('short_answer', {})
                        if isinstance(short_answer, dict):
                            short_answer_text = short_answer.get('answer_text', '')
                        else:
                            short_answer_text = str(short_answer)
                    else:
                        # 字符串结构 - 使用原解析方法
                        root_info = self._parse_question_string(str(root_question))
                        question_text = root_info.get('question', '')
                        document_id = root_info.get('document_id', '')
                        short_answer_text = root_info.get('answer', '')  # 对于根问题，这就是短答案
                    
                    # 添加根问题
                    qa_pairs.append({
                        'Tree_ID': tree_id,
                        'Node_Type': 'Root',
                        'Node_ID': 'root',
                        'Question': question_text,
                        'Answer': short_answer_text,  # 根问题只有短答案，没有长答案
                        'Short_Answer': short_answer_text,
                        'Document_ID': document_id,
                        'Topic': self._extract_topic_from_doc_id(document_id),
                        'TXT_File': f"{document_id}.txt" if document_id else '',
                        'Question_Type': 'factual',
                        'Difficulty': '',
                        'Extension_Type': '',
                        'Depth_Level': 0,
                        'Keywords': '',
                        'Confidence': '',
                        'Verification_Score': ''
                    })
                    
                    # 添加扩展节点
                    nodes = tree_data.get('nodes', {})
                    for node_id, node_data in nodes.items():
                        # 处理ExtensionNode对象
                        ext_info = None  # 初始化ext_info变量
                        if hasattr(node_data, 'question'):
                            # ExtensionNode对象结构
                            ext_question = node_data.question
                            ext_answer = node_data.answer if hasattr(node_data, 'answer') else ''
                            ext_short_answer = getattr(node_data, 'short_answer', ext_answer)
                            ext_type = node_data.extension_type if hasattr(node_data, 'extension_type') else 'unknown'
                            depth_level = node_data.depth_level if hasattr(node_data, 'depth_level') else 0
                            confidence = node_data.confidence if hasattr(node_data, 'confidence') else 0.0
                            verification_score = node_data.verification_score if hasattr(node_data, 'verification_score') else 0.0
                            keywords = node_data.keywords if hasattr(node_data, 'keywords') else []
                            if isinstance(keywords, list):
                                keywords_str = ', '.join(map(str, keywords))
                            else:
                                keywords_str = str(keywords)
                        elif isinstance(node_data, dict):
                            # 字典结构
                            ext_question = node_data.get('question', '')
                            ext_answer = node_data.get('answer', '')
                            
                            # 处理short_answer字典格式
                            short_answer_data = node_data.get('short_answer', ext_answer)
                            if isinstance(short_answer_data, dict) and 'answer_text' in short_answer_data:
                                ext_short_answer = short_answer_data['answer_text']
                            else:
                                ext_short_answer = str(short_answer_data) if short_answer_data else ext_answer
                            
                            ext_type = node_data.get('extension_type', 'unknown')
                            depth_level = node_data.get('depth_level', 0)
                            confidence = node_data.get('confidence', 0.0)
                            verification_score = node_data.get('verification_score', 0.0)
                            keywords = node_data.get('keywords', [])
                            if isinstance(keywords, list):
                                keywords_str = ', '.join(map(str, keywords))
                            else:
                                keywords_str = str(keywords)
                        else:
                            # 字符串结构 - 使用解析方法
                            ext_info = self._parse_extension_string(str(node_data))
                            ext_question = ext_info.get('question', '')
                            ext_answer = ext_info.get('answer', '')
                            ext_short_answer = ext_info.get('short_answer', ext_answer)
                            ext_type = ext_info.get('extension_type', 'unknown')
                            depth_level = ext_info.get('depth_level', 0)
                            confidence = ext_info.get('confidence', 0.0)
                            verification_score = ext_info.get('verification_score', 0.0)
                            keywords_str = ext_info.get('keywords', '')
                        
                        qa_pairs.append({
                            'Tree_ID': tree_id,
                            'Node_Type': 'Extension',
                            'Node_ID': node_id,
                            'Question': ext_question,
                            'Answer': ext_answer,
                            'Short_Answer': ext_short_answer,
                            'Document_ID': document_id,
                            'Topic': self._extract_topic_from_doc_id(document_id),
                            'TXT_File': f"{document_id}.txt" if document_id else '',
                            'Question_Type': f"{ext_type}_extension",
                            'Difficulty': '',
                            'Extension_Type': ext_type,
                            'Depth_Level': depth_level,
                            'Keywords': keywords_str,
                            'Confidence': confidence,
                            'Verification_Score': verification_score
                        })
                
                if qa_pairs:
                    qa_df = pd.DataFrame(qa_pairs)
                    qa_df.to_excel(writer, sheet_name='问答数据', index=False)
                
                # 2. 实验统计摘要
                stats = export_data.get('experiment_metadata', {}).get('statistics', {})
                summary_data = {
                    '统计项目': [
                        '文档总数', '筛选通过文档', '生成问题树数', 
                        '总扩展数', '处理错误数', '完成率(%)'
                    ],
                    '数值': [
                        stats.get('documents_loaded', 0),
                        stats.get('suitable_documents', 0),
                        stats.get('trees_completed', 0),
                        stats.get('extensions_created', 0),
                        stats.get('processing_errors', 0),
                        round(stats.get('trees_completed', 0) / max(stats.get('suitable_documents', 1), 1) * 100, 1)
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='实验统计', index=False)
                
                # 3. 配置信息
                config = export_data.get('experiment_metadata', {}).get('configuration', {})
                config_data = {
                    '配置项': list(config.keys()),
                    '设置值': [str(v) for v in config.values()]
                }
                if config_data['配置项']:
                    config_df = pd.DataFrame(config_data)
                    config_df.to_excel(writer, sheet_name='实验配置', index=False)
                
                # 4. 树结构总览
                tree_summary = []
                for tree_data in export_data.get('question_trees', []):
                    # 安全获取文档ID
                    root_question = tree_data.get('root_question', {})
                    if isinstance(root_question, dict):
                        doc_id = root_question.get('document_id', '')
                    elif hasattr(root_question, 'document_id'):
                        doc_id = root_question.document_id
                    else:
                        doc_id = self._extract_field(str(root_question), 'document_id') if root_question else ''
                    
                    tree_summary.append({
                        '树ID': tree_data['tree_id'],
                        '总节点数': tree_data.get('total_nodes', 0),
                        '最大深度': tree_data.get('max_depth', 0),
                        '扩展数量': tree_data.get('extensions_count', 0),
                        '文档ID': doc_id
                    })
                
                if tree_summary:
                    tree_df = pd.DataFrame(tree_summary)
                    tree_df.to_excel(writer, sheet_name='树结构总览', index=False)
                
                # 5. 轨迹数据 - 增强版
                trajectories_data = export_data.get('trajectories', [])
                if trajectories_data:
                    trajectory_records = []
                    for traj in trajectories_data:
                        if isinstance(traj, dict):
                            # 处理新的轨迹记录格式
                            root_validation = traj.get('root_validation', {})
                            keyword_extraction = traj.get('keyword_extraction', {})
                            extension_steps = traj.get('extension_steps', [])
                            
                            trajectory_records.append({
                                '轨迹ID': traj.get('trajectory_id', ''),
                                '文档ID': traj.get('document_id', ''),
                                '总处理时间': round(traj.get('total_duration', 0.0), 2),
                                '成功率': f"{traj.get('success_rate', 0.0):.1%}",
                                '验证总数': traj.get('total_validations', 0),
                                '成功验证数': traj.get('successful_validations', 0),
                                'Web搜索次数': traj.get('total_web_searches', 0),
                                '最终树深度': traj.get('final_tree_depth', 0),
                                '最终树大小': traj.get('final_tree_size', 0),
                                '平均验证分数': f"{traj.get('average_validation_score', 0.0):.3f}",
                                '关键词层次合规率': f"{traj.get('keyword_hierarchy_compliance', 0.0):.1%}",
                                '快捷路径防止率': f"{traj.get('shortcut_prevention_rate', 0.0):.1%}",
                                '根问题验证分数': f"{root_validation.get('validation_result', {}).get('overall_score', 0.0):.3f}" if root_validation else '',
                                '关键词提取质量': f"{keyword_extraction.get('validation_result', {}).get('keywords_quality', 0.0):.3f}" if keyword_extraction else '',
                                '扩展步骤数': len(extension_steps)
                            })
                        else:
                            # 兼容旧格式
                            trajectory_records.append({
                                '轨迹ID': getattr(traj, 'trajectory_id', ''),
                                '文档ID': getattr(traj, 'document_id', ''),
                                '总处理时间': getattr(traj, 'total_duration', 0),
                                '成功率': f"{getattr(traj, 'success_rate', 0.0):.1%}",
                                '验证总数': getattr(traj, 'total_validations', 0),
                                '成功验证数': getattr(traj, 'successful_validations', 0),
                                'Web搜索次数': 0,
                                '最终树深度': 0,
                                '最终树大小': 0,
                                '平均验证分数': '0.000',
                                '关键词层次合规率': '0.0%',
                                '快捷路径防止率': '0.0%',
                                '根问题验证分数': '',
                                '关键词提取质量': '',
                                '扩展步骤数': 0
                            })
                    
                    if trajectory_records:
                        traj_df = pd.DataFrame(trajectory_records)
                        traj_df.to_excel(writer, sheet_name='轨迹数据', index=False)
            
            logger.info(f"Excel结果已导出: {excel_file}")
            return excel_file
            
        except Exception as e:
            logger.error(f"Excel导出失败: {e}")
            # 不抛出异常，而是继续
            return None
    
    def _format_excel_file(self, excel_file: Path):
        """格式化Excel文件"""
        try:
            wb = openpyxl.load_workbook(excel_file)
            
            # 标题样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # 边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 格式化标题行
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
                
                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # 冻结首行
                ws.freeze_panes = 'A2'
            
            wb.save(excel_file)
            logger.debug(f"Excel格式化完成: {excel_file}")
            
        except Exception as e:
            logger.warning(f"Excel格式化失败: {e}")
    
    def _export_analysis_report(self, export_data: Dict[str, Any], timestamp: str) -> Path:
        """导出分析报告"""
        report_file = self.results_dir / f"analysis_report_{timestamp}.md"
        
        try:
            report_content = self._generate_analysis_report(export_data)
            
            with open(report_file, 'w', encoding='utf-8') as f:
                f.write(report_content)
            
            logger.info(f"分析报告已导出: {report_file}")
            return report_file
            
        except Exception as e:
            logger.error(f"分析报告导出失败: {e}")
            raise
    
    def _generate_analysis_report(self, export_data: Dict[str, Any], timestamp: str) -> str:
        """生成分析报告内容"""
        
        stats = export_data['summary_statistics']
        ext_stats = stats['extension_statistics']
        
        # 计算额外指标
        avg_quality = sum(ext_stats['quality_scores']) / len(ext_stats['quality_scores']) if ext_stats['quality_scores'] else 0
        parallel_ratio = ext_stats['parallel_count'] / max(stats['total_extensions'], 1)
        series_ratio = ext_stats['series_count'] / max(stats['total_extensions'], 1)
        
        report = f"""# Tree Extension Deep Query 实验分析报告

## 📊 实验概览

**生成时间**: {export_data['export_timestamp']}

**核心指标**:
- 📁 处理文档数: {stats['total_documents']}
- 🌳 生成问题树: {stats['total_question_trees']}
- ❓ 总问答对数: {stats['total_qa_pairs']}
- 🔄 总扩展数: {stats['total_extensions']}

## 🔍 扩展分析

### 扩展类型分布
- **Parallel扩展**: {ext_stats['parallel_count']} ({parallel_ratio:.1%})
- **Series扩展**: {ext_stats['series_count']} ({series_ratio:.1%})

### 深度与复杂度
- **最大扩展深度**: {ext_stats['max_depth']}
- **平均分支数**: {ext_stats['avg_branches']:.2f}
- **平均质量分数**: {avg_quality:.3f}

## 📈 质量评估

### 验证通过率
"""
        
        if export_data['trajectories']:
            # 轨迹分析
            trajectories = export_data['trajectories']
            avg_success_rate = sum(t['success_rate'] for t in trajectories) / len(trajectories)
            avg_duration = sum(t['duration'] for t in trajectories) / len(trajectories)
            avg_verification = sum(t['verification_score'] for t in trajectories) / len(trajectories)
            
            report += f"""
- **平均成功率**: {avg_success_rate:.1%}
- **平均验证分数**: {avg_verification:.3f}
- **平均处理时间**: {avg_duration:.2f}秒

### 轨迹统计
- **总轨迹数**: {len(trajectories)}
- **平均步骤数**: {sum(t['total_steps'] for t in trajectories) / len(trajectories):.1f}
- **平均决策点**: {sum(t['decision_points_count'] for t in trajectories) / len(trajectories):.1f}
"""
        
        # 质量分布
        if ext_stats['quality_scores']:
            high_quality = sum(1 for score in ext_stats['quality_scores'] if score >= 0.8)
            medium_quality = sum(1 for score in ext_stats['quality_scores'] if 0.6 <= score < 0.8)
            low_quality = sum(1 for score in ext_stats['quality_scores'] if score < 0.6)
            
            report += f"""
### 质量分数分布
- **高质量** (≥0.8): {high_quality} ({high_quality/len(ext_stats['quality_scores']):.1%})
- **中等质量** (0.6-0.8): {medium_quality} ({medium_quality/len(ext_stats['quality_scores']):.1%})
- **低质量** (<0.6): {low_quality} ({low_quality/len(ext_stats['quality_scores']):.1%})
"""
        
        report += f"""
## 📋 实验配置

| 配置项 | 值 |
|--------|-----|
"""
        
        for key, value in export_data['experiment_metadata'].items():
            report += f"| {key} | {value} |\n"
        
        report += f"""
## 📊 数据统计

### 问答对类型分布
- **根问题**: {stats['total_question_trees']}
- **扩展问题**: {stats['total_extensions']}

### 扩展效率
- **每树平均扩展数**: {stats['total_extensions'] / max(stats['total_question_trees'], 1):.2f}
- **扩展成功率**: {avg_quality:.1%}

## 💡 关键发现

1. **扩展策略**: {"Parallel扩展为主" if parallel_ratio > series_ratio else "Series扩展为主"} ({max(parallel_ratio, series_ratio):.1%})
2. **质量保证**: 平均验证分数 {avg_quality:.3f}
3. **复杂度控制**: 最大深度 {ext_stats['max_depth']} 层，符合设计要求

## 📝 改进建议

1. 根据质量分数分布调整验证阈值
2. 优化{"parallel" if parallel_ratio < 0.3 else "series"}扩展策略的比例
3. 考虑增加{"深度" if ext_stats['max_depth'] < 3 else "广度"}扩展机制

---
*报告生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}*
"""
        
        return report
    
    def export_individual_tree(self, question_tree: Any, output_file: Path):
        """导出单个问题树"""
        try:
            tree_data = {
                'tree_id': question_tree.root_question.document_id,
                'root_question': {
                    'text': question_tree.root_question.text,
                    'answer': question_tree.root_question.short_answer,
                    'type': question_tree.root_question.question_type,
                    'difficulty': question_tree.root_question.difficulty_level,
                    'confidence': question_tree.root_question.confidence
                },
                'extensions': {},
                'metadata': {
                    'total_nodes': question_tree.total_nodes,
                    'max_depth': question_tree.current_depth,
                    'export_timestamp': datetime.now().isoformat()
                }
            }
            
            for node_id, node in question_tree.nodes.items():
                tree_data['extensions'][node_id] = {
                    'question': node.question,
                    'answer': node.answer,
                    'short_answer': getattr(node, 'short_answer', None),
                    'extension_type': node.extension_type,
                    'depth_level': node.depth_level,
                    'parent_node_id': node.parent_node_id,
                    'confidence': node.confidence,
                    'verification_score': node.verification_score,
                    'search_verified': node.search_verified,
                    'keywords': node.keywords,
                    'reasoning': node.reasoning
                }
            
            with open(output_file, 'w', encoding='utf-8') as f:
                json.dump(tree_data, f, indent=2, ensure_ascii=False)
            
            logger.info(f"单个问题树已导出: {output_file}")
            
        except Exception as e:
            logger.error(f"单个问题树导出失败: {e}")
            raise
    
    def get_statistics(self) -> Dict[str, Any]:
        """获取导出统计信息"""
        return self.stats.copy()
    
    def log_export_statistics(self):
        """记录导出统计信息"""
        logger.info("=== 导出系统统计 ===")
        logger.info(f"总导出次数: {self.stats['total_exports']}")
        logger.info(f"JSON导出: {self.stats['json_exports']}")
        logger.info(f"Excel导出: {self.stats['excel_exports']}")
        logger.info(f"失败导出: {self.stats['failed_exports']}")

    def _extract_field(self, text: str, field_name: str) -> Any:
        """从字符串中提取特定字段值"""
        try:
            # 使用正则表达式查找字段名
            match = re.search(f"{field_name}=([^,}}]*)", text)
            if match:
                # 尝试解析为数字或布尔值
                try:
                    return float(match.group(1))
                except ValueError:
                    return match.group(1)
            return None
        except Exception as e:
            logger.warning(f"Failed to extract field '{field_name}' from text: {e}")
            return None
    
    def _parse_question_string(self, question_str: str) -> Dict[str, str]:
        """解析问题字符串，提取关键信息"""
        result = {
            'question': '',
            'answer': '',
            'short_answer': '',
            'document_id': '',
            'question_type': '',
            'difficulty': ''
        }
        
        if not question_str:
            return result
        
        try:
            # 使用正则表达式提取信息
            import re
            
            # 提取 question_text (处理单引号内容)
            question_match = re.search(r"question_text='([^']*(?:''[^']*)*)'", question_str)
            if question_match:
                result['question'] = question_match.group(1).replace("''", "'")
            
            # 提取 ShortAnswer 中的 answer_text (处理单引号内容)
            short_answer_match = re.search(r"ShortAnswer\([^)]*answer_text='([^']*(?:''[^']*)*)'", question_str)
            if short_answer_match:
                result['answer'] = short_answer_match.group(1).replace("''", "'")
            
            # 也尝试从 short_answer= 后面的字典格式提取
            if not result['answer']:
                answer_dict_match = re.search(r"'answer_text':\s*'([^']*(?:''[^']*)*)'", question_str)
                if answer_dict_match:
                    result['answer'] = answer_dict_match.group(1).replace("''", "'")
            
            # 提取 document_id
            doc_id_match = re.search(r"document_id='([^']*)'", question_str)
            if doc_id_match:
                result['document_id'] = doc_id_match.group(1)
            
            # 提取 question_type
            type_match = re.search(r"question_type='([^']*)'", question_str)
            if type_match:
                result['question_type'] = type_match.group(1)
            
            # 提取 difficulty
            diff_match = re.search(r"difficulty='([^']*)'", question_str)
            if diff_match:
                result['difficulty'] = diff_match.group(1)
                
        except Exception as e:
            logger.warning(f"Failed to parse question string: {e}")
        
        return result
    
    def _extract_topic_from_doc_id(self, doc_id: str) -> str:
        """从文档ID提取主题"""
        if not doc_id:
            return 'unknown'
        
        # ClueWeb22格式：clueweb22-en0000-xx-xxxxx
        import re
        topic_match = re.search(r'clueweb22-(en\d+)', doc_id)
        if topic_match:
            return topic_match.group(1)
        return 'unknown'

    def _parse_extension_string(self, ext_str: str) -> Dict[str, str]:
        """解析扩展节点字符串，提取关键信息"""
        result = {
            'question': '',
            'answer': '',
            'short_answer': '',
            'extension_type': '',
            'depth_level': '',
            'keywords': '',
            'confidence': '',
            'verification_score': ''
        }
        
        if not ext_str:
            return result
        
        try:
            import re
            
            # 处理ExtensionNode格式，提取 question (处理复杂引号)
            # 先尝试单引号包围的问题
            question_match = re.search(r"question='([^']*(?:''[^']*)*)'", ext_str)
            if question_match:
                result['question'] = question_match.group(1).replace("''", "'")
            else:
                # 尝试双引号包围的问题
                question_match = re.search(r'question="([^"]*(?:""[^"]*)*)"', ext_str)
                if question_match:
                    result['question'] = question_match.group(1).replace('""', '"')
            
            # 提取 answer (处理复杂引号和长文本)
            # 先尝试单引号包围的答案
            answer_match = re.search(r"answer='([^']*(?:''[^']*)*)'", ext_str)
            if answer_match:
                result['answer'] = answer_match.group(1).replace("''", "'")
            else:
                # 尝试双引号包围的答案
                answer_match = re.search(r'answer="([^"]*(?:""[^"]*)*)"', ext_str)
                if answer_match:
                    result['answer'] = answer_match.group(1).replace('""', '"')
            
            # 提取 short_answer 字典中的 answer_text
            short_answer_dict_match = re.search(r"short_answer=\{[^}]*'answer_text':\s*'([^']*(?:''[^']*)*)'", ext_str)
            if short_answer_dict_match:
                result['short_answer'] = short_answer_dict_match.group(1).replace("''", "'")
            
            # 提取 extension_type
            type_match = re.search(r"extension_type='([^']*)'", ext_str)
            if type_match:
                result['extension_type'] = type_match.group(1)
            
            # 提取 depth_level
            depth_match = re.search(r"depth_level=(\d+)", ext_str)
            if depth_match:
                result['depth_level'] = int(depth_match.group(1))
            
            # 提取 keywords (处理列表格式)
            keywords_match = re.search(r"keywords=\[([^\]]*)\]", ext_str)
            if keywords_match:
                keywords_content = keywords_match.group(1)
                # 提取单引号内的关键词
                keyword_list = re.findall(r"'([^']*)'", keywords_content)
                result['keywords'] = ', '.join(keyword_list)
            
            # 提取 confidence
            conf_match = re.search(r"confidence=([0-9.]+)", ext_str)
            if conf_match:
                result['confidence'] = float(conf_match.group(1))
            
            # 提取 verification_score
            score_match = re.search(r"verification_score=([0-9.]+)", ext_str)
            if score_match:
                result['verification_score'] = float(score_match.group(1))
                
        except Exception as e:
            logger.warning(f"Failed to parse extension string: {e}")
        
        return result


# ====== 测试代码 ======
def test_export_system():
    """测试导出系统"""
    print("=== 测试导出系统 ===")
    
    # 这里需要实际的问题树和轨迹数据来测试
    # 由于依赖其他组件，这里只做基本测试
    
    export_system = ExportSystem()
    
    # 模拟数据
    mock_metadata = {
        'experiment_name': '07_tree_extension_deep_query',
        'version': '1.0.0',
        'max_depth': 3,
        'max_branches': 2
    }
    
    print("导出系统初始化成功")
    print(f"结果目录: {export_system.results_dir}")
    print(f"导出格式: {export_system.export_formats}")
    
    # 显示统计
    stats = export_system.get_statistics()
    print(f"导出统计: {stats}")


if __name__ == "__main__":
    # 配置日志
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    test_export_system() 