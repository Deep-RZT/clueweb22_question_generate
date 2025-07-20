#!/usr/bin/env python3
"""
增量Excel导出器 - 支持处理过程中实时更新Excel文件
防止数据丢失，每处理几个文档就更新一次Excel
"""

import pandas as pd
import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger(__name__)

class IncrementalExcelExporter:
    """增量Excel导出器"""
    
    def __init__(self, excel_file_path: str, update_interval: int = 5):
        self.excel_file = Path(excel_file_path)
        self.update_interval = update_interval  # 每处理N个文档更新一次
        self.processed_count = 0
        
        # 数据缓存
        self.qa_data = []
        self.trajectory_data = []
        self.tree_summary_data = []
        
        # 确保目录存在
        self.excel_file.parent.mkdir(exist_ok=True)
        
        # 初始化Excel文件
        self._initialize_excel_file()
        
        logger.info(f"📊 增量Excel导出器初始化: {self.excel_file}")
    
    def _initialize_excel_file(self):
        """初始化Excel文件结构"""
        try:
            with pd.ExcelWriter(self.excel_file, engine='openpyxl') as writer:
                # 创建空的数据框架
                empty_qa_df = pd.DataFrame(columns=[
                    'Tree_ID', 'Node_Type', 'Node_ID', 'Question', 'Answer', 'Short_Answer',
                    'Document_ID', 'Topic', 'TXT_File', 'Question_Type', 'Difficulty',
                    'Extension_Type', 'Depth_Level', 'Keywords', 'Confidence', 'Verification_Score'
                ])
                empty_qa_df.to_excel(writer, sheet_name='问答数据', index=False)
                
                empty_traj_df = pd.DataFrame(columns=[
                    '轨迹ID', '文档ID', '总处理时间', '成功率', '验证总数', '成功验证数',
                    'Web搜索次数', '最终树深度', '最终树大小', '平均验证分数',
                    '关键词层次合规率', '快捷路径防止率', '根问题验证分数', '关键词提取质量', '扩展步骤数'
                ])
                empty_traj_df.to_excel(writer, sheet_name='轨迹数据', index=False)
                
                empty_summary_df = pd.DataFrame(columns=[
                    '树ID', '总节点数', '最大深度', '扩展数量', '文档ID', '处理时间'
                ])
                empty_summary_df.to_excel(writer, sheet_name='树结构总览', index=False)
                
                # 统计信息表
                stats_df = pd.DataFrame([{
                    '统计项目': '初始化',
                    '数值': f'开始时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}',
                    '备注': '增量更新Excel文件'
                }])
                stats_df.to_excel(writer, sheet_name='实时统计', index=False)
            
            self._format_excel_headers()
            logger.info(f"✅ Excel文件初始化完成")
            
        except Exception as e:
            logger.error(f"初始化Excel文件失败: {e}")
    
    def add_tree_data(self, tree_data: Dict[str, Any], trajectory_data: Optional[Dict[str, Any]] = None):
        """添加单个问答树数据"""
        try:
            # 处理问答数据
            tree_id = tree_data.get('tree_id', f'tree_{len(self.qa_data) + 1}')
            document_id = tree_data.get('document_id', '')
            
            # 根问题数据 - 修复字段提取
            root_question = tree_data.get('root_question', {})
            
            # 初始化默认值
            question_text = ''
            answer_text = 'Unknown'
            confidence = 0.0
            validation_score = 0.0
            
            # 多种格式的解析逻辑 - 强化版本
            if isinstance(root_question, dict):
                # 字典格式 - 优先使用完整字段名
                question_text = (root_question.get('question') or 
                               root_question.get('question_text') or 
                               root_question.get('text') or '')
                answer_text = (root_question.get('answer') or 
                             root_question.get('expected_answer') or 
                             root_question.get('short_answer') or 'Unknown')
                confidence = root_question.get('confidence', 0.0)
                validation_score = root_question.get('validation_score', 0.0)
                
                logger.debug(f"字典格式解析: Q='{question_text[:50]}...', A='{answer_text}'")
                
            elif hasattr(root_question, '__dict__'):
                # 对象格式 - GeneratedQuestion对象
                question_text = getattr(root_question, 'question', getattr(root_question, 'text', ''))
                answer_text = getattr(root_question, 'expected_answer', getattr(root_question, 'answer', 'Unknown'))
                confidence = getattr(root_question, 'confidence', 0.0)
                validation_score = getattr(root_question, 'validation_score', 0.0)
                
                logger.debug(f"对象格式解析: Q='{question_text[:50]}...', A='{answer_text}'")
                
            elif isinstance(root_question, str):
                # 字符串格式 - GeneratedQuestion字符串表示
                question_str = root_question
                logger.debug(f"字符串格式原始: {question_str[:200]}...")
                
                try:
                    import re
                    # 多种正则匹配模式
                    patterns = [
                        (r"question='([^']*)'", r"expected_answer='([^']*)'"),
                        (r'question="([^"]*)"', r'expected_answer="([^"]*)"'),
                        (r"question=([^,)]+)", r"expected_answer=([^,)]+)"),
                    ]
                    
                    parsed = False
                    for q_pattern, a_pattern in patterns:
                        question_match = re.search(q_pattern, question_str)
                        answer_match = re.search(a_pattern, question_str)
                        
                        if question_match and answer_match:
                            question_text = question_match.group(1).strip()
                            answer_text = answer_match.group(1).strip()
                            parsed = True
                            logger.debug(f"正则解析成功: Q='{question_text[:50]}...', A='{answer_text}'")
                            break
                    
                    if not parsed:
                        # 尝试简单分割方法
                        if "expected_answer='" in question_str:
                            answer_start = question_str.find("expected_answer='") + len("expected_answer='")
                            answer_end = question_str.find("'", answer_start)
                            if answer_end > answer_start:
                                answer_text = question_str[answer_start:answer_end]
                                logger.debug(f"分割方法解析答案: '{answer_text}'")
                        
                        if "question='" in question_str:
                            q_start = question_str.find("question='") + len("question='")
                            q_end = question_str.find("'", q_start)
                            if q_end > q_start:
                                question_text = question_str[q_start:q_end]
                                logger.debug(f"分割方法解析问题: '{question_text[:50]}...'")
                    
                    # 提取置信度和验证分数
                    confidence_match = re.search(r"confidence=([0-9.]+)", question_str)
                    validation_match = re.search(r"validation_score=([0-9.]+)", question_str)
                    
                    confidence = float(confidence_match.group(1)) if confidence_match else 0.85
                    validation_score = float(validation_match.group(1)) if validation_match else 0.8
                    
                except Exception as e:
                    logger.warning(f"字符串解析失败: {e}")
                    question_text = question_str[:100] + "..." if len(question_str) > 100 else question_str
                    answer_text = 'Failed to parse'
                    confidence = 0.0
                    validation_score = 0.0
            
            else:
                # 其他格式
                logger.warning(f"未知的root_question格式: {type(root_question)}")
                question_text = str(root_question)[:100] + "..."
                answer_text = 'Unknown format'
            
            # 验证解析结果
            if not answer_text or answer_text in ['Unknown', 'Failed to parse', 'Unknown format', '']:
                logger.warning(f"Root答案解析失败: tree_id={tree_id}, answer='{answer_text}'")
            else:
                logger.info(f"Root答案解析成功: tree_id={tree_id}, answer='{answer_text}'")
            
            # 添加根问题行
            qa_row = {
                'Tree_ID': tree_id,
                'Node_Type': 'Root',
                'Node_ID': 'root',
                'Question': question_text,
                'Answer': answer_text,
                'Short_Answer': answer_text,
                'Document_ID': document_id,
                'Topic': self._extract_topic_from_doc_id(document_id),
                'TXT_File': f"{document_id}.txt" if document_id else '',
                'Question_Type': 'factual',
                'Difficulty': '',
                'Extension_Type': '',
                'Depth_Level': 0,
                'Keywords': '',
                'Confidence': confidence,
                'Verification_Score': validation_score
            }
            self.qa_data.append(qa_row)
            
            # 添加扩展节点
            nodes = tree_data.get('nodes', {})
            for node_id, node_data in nodes.items():
                if isinstance(node_data, dict):
                    ext_question = node_data.get('question', '')
                    ext_answer = node_data.get('answer', '')
                    ext_short_answer = node_data.get('short_answer', ext_answer)
                    ext_type = node_data.get('extension_type', 'unknown')
                    depth_level = node_data.get('depth_level', 1)
                    ext_confidence = node_data.get('confidence', 0.0)
                    
                    ext_row = {
                        'Tree_ID': tree_id,
                        'Node_Type': 'Extension',
                        'Node_ID': node_id,
                        'Question': ext_question,
                        'Answer': ext_answer,
                        'Short_Answer': ext_short_answer,
                        'Document_ID': document_id,
                        'Topic': self._extract_topic_from_doc_id(document_id),
                        'TXT_File': f"{document_id}.txt" if document_id else '',
                        'Question_Type': f"{ext_type}_extension",
                        'Difficulty': '',
                        'Extension_Type': ext_type,
                        'Depth_Level': depth_level,
                        'Keywords': '',
                        'Confidence': ext_confidence,
                        'Verification_Score': 0.0
                    }
                    self.qa_data.append(ext_row)
            
            # 添加轨迹数据 - 确保总是添加轨迹行
            traj_row = {
                '轨迹ID': trajectory_data.get('trajectory_id', '') if trajectory_data else '',
                '文档ID': document_id,
                '总处理时间': round(trajectory_data.get('total_duration', 0.0), 2) if trajectory_data else 0.0,
                '成功率': f"{trajectory_data.get('success_rate', 0.0):.1%}" if trajectory_data else "0.0%",
                '验证总数': trajectory_data.get('total_validations', 0) if trajectory_data else 0,
                '成功验证数': trajectory_data.get('successful_validations', 0) if trajectory_data else 0,
                'Web搜索次数': trajectory_data.get('total_web_searches', 0) if trajectory_data else 0,
                '最终树深度': trajectory_data.get('final_tree_depth', 0) if trajectory_data else 0,
                '最终树大小': trajectory_data.get('final_tree_size', 0) if trajectory_data else 0,
                '平均验证分数': f"{trajectory_data.get('average_validation_score', 0.0):.3f}" if trajectory_data else "0.000",
                '关键词层次合规率': f"{trajectory_data.get('keyword_hierarchy_compliance', 0.0):.1%}" if trajectory_data else "0.0%",
                '快捷路径防止率': f"{trajectory_data.get('shortcut_prevention_rate', 0.0):.1%}" if trajectory_data else "0.0%",
                '根问题验证分数': f"{validation_score:.3f}",
                '关键词提取质量': '0.000',
                '扩展步骤数': len(trajectory_data.get('extension_steps', [])) if trajectory_data else 0
            }
            self.trajectory_data.append(traj_row)
            
            if trajectory_data:
                logger.info(f"   📊 轨迹数据已添加: {trajectory_data.get('trajectory_id', 'N/A')}")
            else:
                logger.warning(f"   ⚠️ 轨迹数据缺失，添加空白行: {document_id}")
            
            # 添加树总览数据
            summary_row = {
                '树ID': tree_id,
                '总节点数': tree_data.get('total_nodes', 0),
                '最大深度': tree_data.get('max_depth', 0),
                '扩展数量': tree_data.get('extensions_count', 0),
                '文档ID': document_id,
                '处理时间': datetime.now().strftime("%H:%M:%S")
            }
            self.tree_summary_data.append(summary_row)
            
            self.processed_count += 1
            
            # 检查是否需要更新Excel
            if self.processed_count % self.update_interval == 0:
                self.update_excel_file()
                logger.info(f"📊 增量更新Excel: 已处理 {self.processed_count} 个文档")
            
        except Exception as e:
            logger.error(f"添加树数据失败: {e}")
    
    def update_excel_file(self):
        """更新Excel文件"""
        try:
            with pd.ExcelWriter(self.excel_file, engine='openpyxl', mode='w') as writer:
                # 问答数据表
                if self.qa_data:
                    qa_df = pd.DataFrame(self.qa_data)
                    qa_df.to_excel(writer, sheet_name='问答数据', index=False)
                
                # 轨迹数据表
                if self.trajectory_data:
                    traj_df = pd.DataFrame(self.trajectory_data)
                    traj_df.to_excel(writer, sheet_name='轨迹数据', index=False)
                
                # 树结构总览表
                if self.tree_summary_data:
                    summary_df = pd.DataFrame(self.tree_summary_data)
                    summary_df.to_excel(writer, sheet_name='树结构总览', index=False)
                
                # 实时统计表
                current_stats = {
                    '统计项目': ['处理时间', '已处理文档', '成功问答对', '平均扩展数', '最新更新时间'],
                    '数值': [
                        f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                        len(self.tree_summary_data),
                        len(self.qa_data),
                        f"{sum(row['扩展数量'] for row in self.tree_summary_data) / max(len(self.tree_summary_data), 1):.1f}",
                        datetime.now().strftime("%H:%M:%S")
                    ],
                    '备注': [
                        '开始处理时间',
                        '包含成功和失败的文档',
                        '根问题 + 扩展问题',
                        '平均每个树的扩展数量',
                        '上次Excel更新时间'
                    ]
                }
                stats_df = pd.DataFrame(current_stats)
                stats_df.to_excel(writer, sheet_name='实时统计', index=False)
            
            self._format_excel_headers()
            
        except Exception as e:
            logger.error(f"更新Excel文件失败: {e}")
    
    def final_update(self):
        """最终更新Excel文件"""
        try:
            self.update_excel_file()
            logger.info(f"✅ Excel文件最终更新完成: {self.excel_file}")
            logger.info(f"📊 最终统计: {len(self.tree_summary_data)} 个树, {len(self.qa_data)} 个问答对, {len(self.trajectory_data)} 个轨迹")
        except Exception as e:
            logger.error(f"最终更新Excel失败: {e}")
    
    def _extract_topic_from_doc_id(self, doc_id: str) -> str:
        """从文档ID提取topic"""
        if 'en00' in doc_id:
            # 例如: clueweb22-en0028-01-00001 -> en0028
            parts = doc_id.split('-')
            for part in parts:
                if part.startswith('en00'):
                    return part
        return 'unknown'
    
    def _format_excel_headers(self):
        """格式化Excel表头"""
        try:
            wb = openpyxl.load_workbook(self.excel_file)
            
            # 表头样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 格式化第一行（表头）
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                
                # 冻结首行
                ws.freeze_panes = 'A2'
            
            wb.save(self.excel_file)
            
        except Exception as e:
            logger.warning(f"格式化Excel表头失败: {e}")
    
    def get_stats(self) -> Dict[str, int]:
        """获取当前统计信息"""
        return {
            'processed_documents': len(self.tree_summary_data),
            'total_qa_pairs': len(self.qa_data),
            'total_trajectories': len(self.trajectory_data),
            'excel_updates': self.processed_count // self.update_interval
        } 